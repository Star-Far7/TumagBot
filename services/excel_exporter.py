import io
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_invoice_excel(items: List[Dict], invoice_date: Optional[str] = None) -> bytes:
    """
    Собирает Excel-файл из подтверждённых позиций накладной.
    Лист 1 «Для Umag»  — чистые данные без заголовков, для импорта.
    Лист 2 «Детали»    — полная информативная таблица с форматированием.
    Лист 3 «Изменения цен» — только если есть позиции с изменением цены.
    """
    wb = openpyxl.Workbook()

    # Лист 1 — импорт в Umag (без заголовков, без итогов)
    ws1 = wb.active
    _fill_umag_sheet(ws1, items)

    # Лист 2 — детальная информативная таблица
    ws2 = wb.create_sheet("Детали")
    _fill_detail_sheet(ws2, items, invoice_date)

    # Лист 3 — изменения цен (опционально)
    price_changes = [
        i for i in items
        if i.get("old_price") and i.get("price")
        and i["old_price"] > 0
        and abs(i["price"] - i["old_price"]) > 0.001
        and i.get("match_type") != "skipped"
    ]
    if price_changes:
        ws3 = wb.create_sheet("Изменения цен")
        _fill_price_sheet(ws3, price_changes)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Стили ─────────────────────────────────────────────────────────────────────

_BLUE_HDR   = PatternFill("solid", fgColor="1F4E79")
_RED_ROW    = PatternFill("solid", fgColor="FFDCE0")
_GREEN_ROW  = PatternFill("solid", fgColor="D6F4D6")
_GREY_ROW   = PatternFill("solid", fgColor="EEEEEE")
_WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
_BOLD       = Font(bold=True, size=11)
_CENTER     = Alignment(horizontal="center", vertical="center")
_RIGHT      = Alignment(horizontal="right",  vertical="center")
_LEFT       = Alignment(horizontal="left",   vertical="center")
_NUM_FMT    = "#,##0.00"
_SIGN_FMT   = "+#,##0.00;-#,##0.00"


def _hdr(ws, col: int, row: int, value: str):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _WHITE_BOLD
    c.fill = _BLUE_HDR
    c.alignment = _CENTER


# ── Лист 1: «Для Umag» ────────────────────────────────────────────────────────

def _fill_umag_sheet(ws, items: List[Dict]):
    """
    Чистая таблица для импорта в Umag:
    штрихкод | наименование | ед.изм. | кол-во | цена закупа
    Без заголовков, без дат, без итогов. Пропущенные позиции не включаются.
    """
    ws.title = "Для Umag"

    col_widths = [18, 46, 8, 8, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for item in items:
        if item.get("match_type") == "skipped":
            continue

        barcode    = item.get("product_barcode") or item.get("supplier_barcode") or ""
        name       = item.get("product_name") or item.get("supplier_name") or "—"
        unit       = item.get("unit") or "шт"
        qty        = item.get("quantity") or 1
        unit_price = item.get("price") or 0.0

        row_data = [barcode, name, unit, qty, unit_price]
        ws.append(row_data)

        r = ws.max_row
        ws.cell(r, 1).alignment = _LEFT
        ws.cell(r, 2).alignment = _LEFT
        ws.cell(r, 3).alignment = _CENTER
        ws.cell(r, 4).alignment = _RIGHT
        ws.cell(r, 4).number_format = "#,##0.##"
        ws.cell(r, 5).alignment = _RIGHT
        ws.cell(r, 5).number_format = _NUM_FMT


# ── Лист 2: «Детали» ──────────────────────────────────────────────────────────

def _fill_detail_sheet(ws, items: List[Dict], invoice_date: Optional[str]):
    """
    Информативная таблица с заголовками, цветовой подсветкой и итогами.
    """
    ws.title = "Детали"

    title_date = invoice_date or datetime.now().strftime("%d.%m.%Y")
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Накладная от {title_date}"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = _CENTER

    headers = [
        ("Штрихкод",       16),
        ("Наименование",   42),
        ("Из накладной",   42),
        ("Категория",      18),
        ("Ед.изм.",         8),
        ("Кол-во",          8),
        ("Цена закупа",    13),
        ("Сумма",          13),
        ("Стар. цена",     12),
        ("Изм. %",         10),
    ]
    for col, (header, width) in enumerate(headers, 1):
        _hdr(ws, col, 2, header)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 22

    total = 0.0
    for r, item in enumerate(items, 3):
        barcode       = item.get("product_barcode") or item.get("supplier_barcode") or ""
        product_name  = item.get("product_name") or item.get("supplier_name") or "—"
        supplier_name = item.get("supplier_name") or ""
        category      = item.get("category") or ""
        unit          = item.get("unit") or "шт"
        qty           = item.get("quantity") or 1
        price         = item.get("price") or 0.0
        old_price     = item.get("old_price")
        match_type    = item.get("match_type") or "manual"
        amount        = qty * price
        total        += amount

        row_vals = [
            barcode,
            product_name,
            supplier_name if supplier_name != product_name else "",
            category,
            unit,
            qty,
            price,
            amount,
            old_price if old_price else "",
            "",  # change %
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = _LEFT
            if col in (6, 7, 8, 9):
                cell.number_format = _NUM_FMT
                cell.alignment = _RIGHT
            if col == 6:
                cell.number_format = "#,##0.##"

        # Подсветка изменения цены
        fill = None
        if match_type == "skipped":
            fill = _GREY_ROW
        elif match_type == "new_product":
            pass  # без подсветки
        elif old_price and old_price > 0 and price:
            pct = (price - old_price) / old_price * 100
            change_cell = ws.cell(row=r, column=10, value=f"{pct:+.1f}%")
            change_cell.alignment = _CENTER
            fill = _RED_ROW if pct > 0 else _GREEN_ROW

        if fill:
            for col in range(1, 11):
                ws.cell(row=r, column=col).fill = fill

    # Итого
    total_row = len(items) + 3
    ws.cell(row=total_row, column=7, value="ИТОГО:").font = _BOLD
    tot_cell = ws.cell(row=total_row, column=8, value=total)
    tot_cell.font = _BOLD
    tot_cell.number_format = _NUM_FMT
    tot_cell.alignment = _RIGHT

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:J{len(items) + 2}"


# ── Лист 3: «Изменения цен» ───────────────────────────────────────────────────

def _fill_price_sheet(ws, changes: List[Dict]):
    ws.title = "Изменения цен"
    headers = [
        ("Наименование", 42),
        ("Стар. цена",   13),
        ("Нов. цена",    13),
        ("Изменение",    13),
        ("Изм. %",       10),
    ]
    for col, (header, width) in enumerate(headers, 1):
        _hdr(ws, col, 1, header)
        ws.column_dimensions[get_column_letter(col)].width = width

    sorted_changes = sorted(
        changes,
        key=lambda x: -(x["price"] - x["old_price"]) / x["old_price"],
    )
    for r, item in enumerate(sorted_changes, 2):
        pct = (item["price"] - item["old_price"]) / item["old_price"] * 100
        fill = _RED_ROW if pct > 0 else _GREEN_ROW

        vals = [
            item.get("product_name") or item.get("supplier_name"),
            item["old_price"],
            item["price"],
            item["price"] - item["old_price"],
            f"{pct:+.1f}%",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            if col in (2, 3, 4):
                cell.number_format = _NUM_FMT if col < 4 else _SIGN_FMT
                cell.alignment = _RIGHT
            else:
                cell.alignment = _LEFT if col == 1 else _CENTER
